import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import csv
import fitz  # PyMuPDF
import os
from openpyxl import Workbook, load_workbook
from ttkbootstrap import Style
import threading
from PIL import Image, ImageTk

class CertificateValidator:
    def __init__(self, root):
        self.root = root
        self.style = Style(theme='morph')
        self.root.title("Certificate Validator Pro")
        self.root.geometry("1600x900")
        
        # Initialize variables
        self.csv_data = {}
        self.certificate_folder = ""
        self.stats = {'scanned': 0, 'errors': 0, 'valid': 0}
        self.current_certificate = None
        self.current_csv_row = None

        # Setup UI
        self.create_layout()
        self.style.configure('TLabel', font=('Segoe UI', 10))
        self.style.configure('TButton', font=('Segoe UI', 10))

    def create_layout(self):
        # Main container with split panes
        main_pane = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True)

        # Left Panel (CSV Data)
        left_frame = ttk.Frame(main_pane, width=600)
        main_pane.add(left_frame)
        self.create_left_panel(left_frame)

        # Right Panel (Certificate Preview)
        right_frame = ttk.Frame(main_pane, width=900)
        main_pane.add(right_frame)
        self.create_right_panel(right_frame)

    def create_left_panel(self, parent):
        # CSV Data Treeview
        csv_frame = ttk.LabelFrame(parent, text=" Reference Data ", padding=10)
        csv_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.tree = ttk.Treeview(csv_frame, columns=('RefNo', 'Name', 'School'), show='headings')
        self.tree.heading('RefNo', text='Reference Number')
        self.tree.heading('Name', text='Full Name')
        self.tree.heading('School', text='School Name')
        self.tree.column('RefNo', width=150)
        self.tree.column('Name', width=250)
        self.tree.column('School', width=200)
        
        vsb = ttk.Scrollbar(csv_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(csv_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        csv_frame.grid_rowconfigure(0, weight=1)
        csv_frame.grid_columnconfigure(0, weight=1)

        # Control Panel
        control_frame = ttk.Frame(parent)
        control_frame.pack(pady=10, fill=tk.X)
        
        ttk.Button(control_frame, 
                 text="Load CSV", 
                 command=self.load_csv).pack(side=tk.LEFT, padx=5)
        ttk.Button(control_frame, 
                 text="Select PDF Folder", 
                 command=self.select_folder).pack(side=tk.LEFT, padx=5)
        self.validate_btn = ttk.Button(control_frame, 
                                     text="Start Validation", 
                                     command=self.start_validation, 
                                     state='disabled')
        self.validate_btn.pack(side=tk.LEFT, padx=5)

    def create_right_panel(self, parent):
        # Certificate Preview (A4 Size)
        cert_frame = ttk.LabelFrame(parent, text=" Certificate Preview ", padding=10)
        cert_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.canvas = tk.Canvas(cert_frame, bg='white', width=794, height=1123)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        
        # Validation Status
        self.status_label = ttk.Label(cert_frame, 
                                    text="Status: Ready",
                                    font=('Segoe UI', 12))
        self.status_label.pack(pady=10)

    def load_csv(self):
        file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
        if not file_path:
            return
            
        try:
            with open(file_path, 'r') as f:
                reader = csv.DictReader(f)
                required_columns = ['Reference Number', 'First Name', 'Last Name', 'School Name']
                
                for col in required_columns:
                    if col not in reader.fieldnames:
                        messagebox.showerror("Error", f"Missing column: {col}")
                        return
                
                self.csv_data.clear()
                self.tree.delete(*self.tree.get_children())
                
                for row in reader:
                    ref_number = row['Reference Number'].strip()
                    full_name = f"{row['First Name']} {row['Last Name']}"
                    self.csv_data[ref_number] = {
                        'full_name': full_name.lower(),
                        'school': row['School Name'].lower()
                    }
                    self.tree.insert('', 'end', 
                                  values=(ref_number, full_name, row['School Name']))
                
                messagebox.showinfo("Success", f"Loaded {len(self.csv_data)} records!")
                self.validate_btn.config(state='normal')
                
        except Exception as e:
            messagebox.showerror("Error", f"CSV Error: {str(e)}")

    def select_folder(self):
        self.certificate_folder = filedialog.askdirectory()
        if self.certificate_folder:
            messagebox.showinfo("Folder Selected", f"Selected: {self.certificate_folder}")

    def start_validation(self):
        if not self.csv_data:
            messagebox.showerror("Error", "Please load CSV file first!")
            return
        if not self.certificate_folder:
            messagebox.showerror("Error", "Please select PDF folder first!")
            return
            
        # Reset UI and counters
        self.tree.selection_remove(self.tree.selection())
        self.canvas.delete("all")
        self.status_label.config(text="Status: Validating...")
        self.stats = {'scanned': 0, 'errors': 0, 'valid': 0}
        
        # Start validation thread
        threading.Thread(target=self.run_validation, daemon=True).start()

    def run_validation(self):
        error_list = []
        
        for filename in os.listdir(self.certificate_folder):
            if filename.endswith(".pdf"):
                try:
                    self.stats['scanned'] += 1
                    ref_number = os.path.splitext(filename)[0].strip()
                    pdf_path = os.path.join(self.certificate_folder, filename)
                    
                    # Update UI
                    self.root.after(0, self.update_current_certificate, filename, pdf_path)
                    self.root.after(0, self.highlight_csv_row, ref_number)
                    
                    # Validate certificate
                    validation_result = self.validate_certificate(ref_number, pdf_path)
                    
                    if validation_result['is_valid']:
                        self.stats['valid'] += 1
                    else:
                        self.stats['errors'] += 1
                        error_list.append({
                            'filename': filename,
                            'reference': ref_number,
                            'errors': validation_result['errors']
                        })
                    
                    # Update stats
                    self.root.after(0, self.update_stats)
                    
                except Exception as e:
                    self.log_to_console(f"Error processing {filename}: {str(e)}")
        
        self.save_to_excel(error_list)
        self.root.after(0, lambda: self.status_label.config(text="Status: Validation Complete"))

    def update_current_certificate(self, filename, pdf_path):
        try:
            doc = fitz.open(pdf_path)
            page = doc.load_page(0)
            pix = page.get_pixmap(dpi=150)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img.thumbnail((794, 1123))  # Maintain aspect ratio
            
            self.current_certificate = ImageTk.PhotoImage(img)
            self.canvas.create_image(0, 0, anchor=tk.NW, image=self.current_certificate)
            self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))
            
        except Exception as e:
            self.canvas.create_text(100, 100, text=f"Preview Error: {str(e)}", fill="red")

    def highlight_csv_row(self, ref_number):
        for child in self.tree.get_children():
            if self.tree.item(child)['values'][0] == ref_number:
                self.tree.selection_set(child)
                self.tree.see(child)
                break

    def validate_certificate(self, ref_number, pdf_path):
        result = {'is_valid': True, 'errors': []}
        
        if ref_number not in self.csv_data:
            result['errors'].append("Reference mismatch")
            result['is_valid'] = False
            return result
            
        try:
            doc = fitz.open(pdf_path)
            text = "".join([page.get_text() for page in doc]).lower()
            expected = self.csv_data[ref_number]
            
            if expected['full_name'] not in text:
                result['errors'].append("Name mismatch")
                result['is_valid'] = False
                
            if expected['school'] not in text:
                result['errors'].append("School mismatch")
                result['is_valid'] = False
                
        except Exception as e:
            result['errors'].append(f"PDF Error: {str(e)}")
            result['is_valid'] = False
            
        return result

    def update_stats(self):
        self.status_label.config(
            text=f"Scanned: {self.stats['scanned']} | "
                 f"Valid: {self.stats['valid']} | "
                 f"Errors: {self.stats['errors']}"
        )

    def save_to_excel(self, error_list):
        try:
            excel_path = "validation_errors.xlsx"
            
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(["Filename", "Reference Number", "Errors"])
            
            for error in error_list:
                ws.append([
                    error['filename'],
                    error['reference'],
                    ", ".join(error['errors'])
                ])
                
            wb.save(excel_path)
            self.log_to_console(f"Error report saved to {excel_path}")
            
        except Exception as e:
            self.log_to_console(f"Excel Save Error: {str(e)}")

    def log_to_console(self, message):
        self.root.after(0, lambda: self.status_label.config(text=message))

if __name__ == "__main__":
    root = tk.Tk()
    app = CertificateValidator(root)
    root.mainloop()